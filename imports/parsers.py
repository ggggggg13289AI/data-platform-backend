"""
File parsing utilities for CSV and Excel files.

Provides:
- CSV/Excel file parsing with header detection
- Column type detection
- Field mapping suggestions
"""

import csv
import io
import logging
import re
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# Field mappings for auto-suggestion
REPORT_FIELD_MAPPINGS = {
    "uid": ["uid", "id", "report_uid", "unique_id"],
    "title": ["title", "report_title", "name", "subject"],
    "content": ["content", "report_content", "body", "text", "description"],
    "report_type": ["report_type", "type", "format", "category"],
    "report_id": ["report_id", "reportid", "report_no", "report_number"],
    "chr_no": ["chr_no", "chrno", "char_no", "character_no"],
    "mod": ["mod", "modality", "mode"],
    "report_date": ["report_date", "date", "created_date", "exam_date"],
    "source_url": ["source_url", "url", "source", "link"],
}

STUDY_FIELD_MAPPINGS = {
    # Primary identifiers
    "exam_id": ["exam_id", "examid", "examination_id", "study_id", "accession_number", "申請單號"],
    "medical_record_no": ["medical_record_no", "mrn", "patient_id", "patient_no", "病歷號"],
    "application_order_no": ["application_order_no", "order_no", "application_no"],
    # Patient info
    "patient_name": ["patient_name", "name", "patient", "full_name", "姓名"],
    "patient_gender": ["patient_gender", "gender", "sex", "性別"],
    "patient_birth_date": ["patient_birth_date", "birth_date", "dob", "date_of_birth", "生日"],
    "patient_age": ["patient_age", "age", "年齡"],
    # Exam details
    "exam_status": ["exam_status", "status", "study_status", "狀態"],
    "exam_source": ["exam_source", "source", "modality", "mod", "imaging_modality", "來源"],
    "exam_item": ["exam_item", "item", "procedure", "study_description", "檢查項目"],
    "exam_description": ["exam_description", "description", "exam_desc", "檢查描述"],
    "exam_room": ["exam_room", "room", "location"],
    "exam_equipment": ["exam_equipment", "equipment", "scanner", "儀器"],
    "equipment_type": ["equipment_type", "equip_type", "device_type"],
    # Temporal fields
    "order_datetime": ["order_datetime", "order_date", "order_time", "開單日期", "開單日期/時間"],
    "check_in_datetime": ["check_in_datetime", "checkin_date", "arrival_time", "報到時間"],
    "report_certification_datetime": [
        "report_certification_datetime",
        "certification_date",
        "report_date",
        "檢查結束時間",
    ],
    # Authorization
    "certified_physician": ["certified_physician", "physician", "doctor", "radiologist"],
}


def detect_column_type(values: list[Any]) -> str:
    """
    Detect the data type of a column based on sample values.

    Args:
        values: List of sample values from the column

    Returns:
        Detected type: 'number', 'date', 'boolean', or 'string'
    """
    if not values:
        return "string"

    # Filter out None/empty values
    non_empty = [v for v in values if v is not None and str(v).strip()]
    if not non_empty:
        return "string"

    # Check for boolean
    bool_values = {"true", "false", "yes", "no", "1", "0", "t", "f", "y", "n"}
    if all(str(v).lower().strip() in bool_values for v in non_empty):
        return "boolean"

    # Check for number
    number_count = 0
    for v in non_empty:
        try:
            float(str(v).replace(",", ""))
            number_count += 1
        except (ValueError, TypeError):
            pass
    if number_count == len(non_empty):
        return "number"

    # Check for date patterns
    date_patterns = [
        r"\d{4}-\d{2}-\d{2}",  # YYYY-MM-DD
        r"\d{2}/\d{2}/\d{4}",  # MM/DD/YYYY or DD/MM/YYYY
        r"\d{4}/\d{2}/\d{2}",  # YYYY/MM/DD
        r"\d{2}-\d{2}-\d{4}",  # MM-DD-YYYY or DD-MM-YYYY
    ]
    date_count = 0
    for v in non_empty:
        str_v = str(v)
        for pattern in date_patterns:
            if re.match(pattern, str_v):
                date_count += 1
                break
    if date_count >= len(non_empty) * 0.8:  # 80% threshold
        return "date"

    return "string"


def suggest_field_mapping(column_name: str, target_type: str) -> str | None:
    """
    Suggest a target field based on column name.

    Args:
        column_name: The source column name
        target_type: 'report' or 'study'

    Returns:
        Suggested target field name or None
    """
    mappings = REPORT_FIELD_MAPPINGS if target_type == "report" else STUDY_FIELD_MAPPINGS
    column_lower = column_name.lower().strip().replace(" ", "_")

    for field, aliases in mappings.items():
        if column_lower in aliases or column_lower == field:
            return field
        # Partial match
        for alias in aliases:
            if alias in column_lower or column_lower in alias:
                return field

    return None


def get_target_fields(target_type: str) -> list[str]:
    """Get list of available target fields for a target type."""
    if target_type == "report":
        return list(REPORT_FIELD_MAPPINGS.keys())
    elif target_type == "study":
        return list(STUDY_FIELD_MAPPINGS.keys())
    return []


def parse_csv(file_path: str, preview_rows: int = 10) -> dict[str, Any]:
    """
    Parse a CSV file and return headers and preview data.

    Args:
        file_path: Path to the CSV file
        preview_rows: Number of rows to preview

    Returns:
        Dict with columns, preview_rows, and total_rows
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    # Try different encodings
    encodings = ["utf-8", "utf-8-sig", "latin-1", "cp1252", "gbk", "big5"]
    content = None

    for encoding in encodings:
        try:
            with open(path, encoding=encoding) as f:
                content = f.read()
            break
        except UnicodeDecodeError:
            continue

    if content is None:
        raise ValueError("Unable to decode file with supported encodings")

    # Parse CSV
    reader = csv.DictReader(io.StringIO(content))
    headers = reader.fieldnames or []

    if not headers:
        raise ValueError("CSV file has no headers")

    # Read all rows for counting and preview
    all_rows = list(reader)
    total_rows = len(all_rows)
    preview_data = all_rows[:preview_rows]

    # Collect sample values for each column
    columns = []
    for header in headers:
        sample_values = [row.get(header, "") for row in preview_data]
        detected_type = detect_column_type(sample_values)

        columns.append(
            {
                "name": header,
                "detected_type": detected_type,
                "sample_values": [str(v)[:100] for v in sample_values[:5]],  # First 5, truncated
                "suggested_field": None,  # Will be filled later based on target type
            }
        )

    return {
        "columns": columns,
        "preview_rows": preview_data,
        "total_rows": total_rows,
        "sheet_names": None,
    }


def parse_excel(
    file_path: str, sheet_name: str | None = None, preview_rows: int = 10
) -> dict[str, Any]:
    """
    Parse an Excel file and return headers and preview data.

    Args:
        file_path: Path to the Excel file
        sheet_name: Optional sheet name (defaults to first sheet)
        preview_rows: Number of rows to preview

    Returns:
        Dict with columns, preview_rows, total_rows, and sheet_names
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel file support. Install with: pip install openpyxl"
        )

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    # Load workbook
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    # Select sheet
    if sheet_name:
        if sheet_name not in sheet_names:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {sheet_names}")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # Read data
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel sheet is empty")

    headers = [str(h) if h is not None else f"Column_{i}" for i, h in enumerate(rows[0])]
    data_rows = rows[1:]  # Skip header row
    total_rows = len(data_rows)

    # Convert to dict format
    preview_data = []
    for row in data_rows[:preview_rows]:
        row_dict = {}
        for i, value in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = value
        preview_data.append(row_dict)

    # Collect sample values for each column
    columns = []
    for i, header in enumerate(headers):
        sample_values = [row[i] if i < len(row) else None for row in data_rows[:preview_rows]]
        detected_type = detect_column_type(sample_values)

        columns.append(
            {
                "name": header,
                "detected_type": detected_type,
                "sample_values": [str(v)[:100] if v is not None else "" for v in sample_values[:5]],
                "suggested_field": None,
            }
        )

    wb.close()

    return {
        "columns": columns,
        "preview_rows": preview_data,
        "total_rows": total_rows,
        "sheet_names": sheet_names,
    }


def parse_file(
    file_path: str, sheet_name: str | None = None, preview_rows: int = 10
) -> dict[str, Any]:
    """
    Parse a file (CSV or Excel) and return preview data.

    Args:
        file_path: Path to the file
        sheet_name: Optional sheet name for Excel files
        preview_rows: Number of rows to preview

    Returns:
        Dict with columns, preview_rows, total_rows, and optionally sheet_names
    """
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix == ".csv":
        return parse_csv(file_path, preview_rows)
    elif suffix in [".xlsx", ".xls"]:
        return parse_excel(file_path, sheet_name, preview_rows)
    else:
        raise ValueError(
            f"Unsupported file type: {suffix}. Only CSV and Excel (.xlsx) files are supported."
        )


def read_csv_rows(file_path: str) -> list[dict[str, Any]]:
    """
    Read all rows from a CSV file.

    Args:
        file_path: Path to the CSV file

    Returns:
        List of row dictionaries
    """
    path = Path(file_path)
    encodings = ["utf-8", "utf-8-sig", "latin-1", "cp1252", "gbk", "big5"]

    for encoding in encodings:
        try:
            with open(path, encoding=encoding) as f:
                reader = csv.DictReader(f)
                return list(reader)
        except UnicodeDecodeError:
            continue

    raise ValueError("Unable to decode file with supported encodings")


def read_excel_rows(file_path: str, sheet_name: str | None = None) -> list[dict[str, Any]]:
    """
    Read all rows from an Excel file.

    Args:
        file_path: Path to the Excel file
        sheet_name: Optional sheet name

    Returns:
        List of row dictionaries
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel file support")

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h) if h is not None else f"Column_{i}" for i, h in enumerate(rows[0])]
    data = []

    for row in rows[1:]:
        row_dict = {}
        for i, value in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = value
        data.append(row_dict)

    wb.close()
    return data


def read_file_rows(file_path: str, sheet_name: str | None = None) -> list[dict[str, Any]]:
    """
    Read all rows from a file (CSV or Excel).

    Args:
        file_path: Path to the file
        sheet_name: Optional sheet name for Excel files

    Returns:
        List of row dictionaries
    """
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix == ".csv":
        return read_csv_rows(file_path)
    elif suffix in [".xlsx", ".xls"]:
        return read_excel_rows(file_path, sheet_name)
    else:
        raise ValueError(f"Unsupported file type: {suffix}")
